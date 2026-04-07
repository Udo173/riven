# gantt_generator.py
# PM Tool - Gantt Chart Generator
# Claude Architecture | OpenCode Implementation
# Path: C:\Users\USass\Billi_Bilder\PM_Tool\core\

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

C_HEADER_BG = "1F497D"
C_HEADER_FG = "FFFFFF"
C_TASK_BG = "F0F4F9"
C_WEEKEND = "E8E8E8"
C_BAR_ACTIVE = "2E75B6"
C_BAR_DONE = "70AD47"
C_BAR_LATE = "FF0000"
C_MILESTONE = "FFC000"
C_ALT_ROW = "DCE6F1"


def col_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def create_gantt(data: dict, output_path: str = None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"

    project_name = data.get("project_name", "Projekt")
    project_mgr = data.get("project_manager", "-")
    tasks = data.get("tasks", [])

    all_dates = []
    for t in tasks:
        try:
            all_dates.append(date.fromisoformat(t["start"]))
            all_dates.append(date.fromisoformat(t["end"]))
        except Exception:
            pass

    if not all_dates:
        chart_start = date.today()
        chart_end = chart_start + timedelta(days=30)
    else:
        chart_start = min(all_dates) - timedelta(days=2)
        chart_end = max(all_dates) + timedelta(days=2)

    total_days = (chart_end - chart_start).days + 1

    INFO_COLS = 7
    DATE_START = INFO_COLS + 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 14

    for i in range(total_days):
        col_letter = get_column_letter(DATE_START + i)
        ws.column_dimensions[col_letter].width = 2.8

    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A1:{get_column_letter(INFO_COLS + total_days)}1")
    c = ws["A1"]
    c.value = f"GANTT CHART - {project_name.upper()}"
    c.font = Font(bold=True, size=14, color=C_HEADER_FG)
    c.fill = col_fill(C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 18
    ws.merge_cells(f"A2:{get_column_letter(INFO_COLS + total_days)}2")
    c2 = ws["A2"]
    c2.value = (
        f"Projektleiter: {project_mgr}   |   "
        f"Erstellt: {date.today().strftime('%d.%m.%Y')}   |   "
        f"Zeitraum: {chart_start.strftime('%d.%m.%Y')} - "
        f"{chart_end.strftime('%d.%m.%Y')}"
    )
    c2.font = Font(size=9, color="444444")
    c2.fill = col_fill("DCE6F1")
    c2.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 22
    headers = ["#", "Aufgabe", "Verantwortlich", "Start", "Ende", "Tage", "Status"]
    for col_i, hdr in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_i)
        c.value = hdr
        c.font = Font(bold=True, size=10, color=C_HEADER_FG)
        c.fill = col_fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    for day_i in range(total_days):
        d = chart_start + timedelta(days=day_i)
        col = DATE_START + day_i
        c = ws.cell(row=3, column=col)
        day_names = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
        c.value = day_names[d.weekday()]
        c.font = Font(bold=True, size=7, color=C_HEADER_FG)
        c.fill = col_fill("2E75B6" if d.weekday() < 5 else "888888")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    ws.row_dimensions[4].height = 16
    ws.merge_cells("A4:G4")
    ws["A4"].value = ""

    for day_i in range(total_days):
        d = chart_start + timedelta(days=day_i)
        col = DATE_START + day_i
        c = ws.cell(row=4, column=col)
        c.value = d.day
        c.font = Font(size=7, color="444444" if d.weekday() < 5 else "888888")
        c.fill = col_fill("FFFFFF" if d.weekday() < 5 else C_WEEKEND)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()

    today = date.today()

    for task_i, task in enumerate(tasks):
        row = 5 + task_i
        ws.row_dimensions[row].height = 20

        try:
            t_start = date.fromisoformat(task["start"])
            t_end = date.fromisoformat(task["end"])
            duration = (t_end - t_start).days + 1
        except Exception:
            t_start = t_end = today
            duration = 1

        status = task.get("status", "Offen")
        is_done = status.lower() in ["fertig", "done", "abgeschlossen"]
        is_late = (not is_done) and t_end < today

        row_bg = C_ALT_ROW if task_i % 2 == 0 else "FFFFFF"

        info_values = [
            task_i + 1,
            task.get("name", "-"),
            task.get("owner", "-"),
            t_start.strftime("%d.%m.%Y"),
            t_end.strftime("%d.%m.%Y"),
            duration,
            status,
        ]
        for col_i, val in enumerate(info_values, 1):
            c = ws.cell(row=row, column=col_i)
            c.value = val
            c.font = Font(size=9)
            c.fill = col_fill(row_bg)
            c.alignment = Alignment(
                vertical="center",
                horizontal="center" if col_i in [1, 4, 5, 6] else "left",
            )
            c.border = thin_border()

            if col_i == 7:
                if is_done:
                    c.font = Font(size=9, color=C_BAR_DONE, bold=True)
                elif is_late:
                    c.font = Font(size=9, color=C_BAR_LATE, bold=True)

        for day_i in range(total_days):
            d = chart_start + timedelta(days=day_i)
            col = DATE_START + day_i
            c = ws.cell(row=row, column=col)
            c.border = thin_border()

            in_task = t_start <= d <= t_end

            if in_task:
                if is_done:
                    bar_color = C_BAR_DONE
                elif is_late:
                    bar_color = C_BAR_LATE
                elif task.get("milestone", False):
                    bar_color = C_MILESTONE
                else:
                    bar_color = C_BAR_ACTIVE
                c.fill = col_fill(bar_color)
                c.value = "X"
                c.font = Font(size=6, color=bar_color)
            elif d.weekday() >= 5:
                c.fill = col_fill(C_WEEKEND)
            else:
                c.fill = col_fill(row_bg)

    today_offset = (today - chart_start).days
    if 0 <= today_offset < total_days:
        col = DATE_START + today_offset
        c3 = ws.cell(row=3, column=col)
        c3.fill = col_fill("FF0000")
        c3.font = Font(bold=True, size=7, color="FFFFFF")
        c4 = ws.cell(row=4, column=col)
        c4.font = Font(size=7, color="FF0000", bold=True)

    legend_row = 6 + len(tasks)
    ws.row_dimensions[legend_row].height = 16
    legend_items = [
        (C_BAR_ACTIVE, "In Arbeit"),
        (C_BAR_DONE, "Fertig"),
        (C_BAR_LATE, "Ueberfaellig"),
        (C_MILESTONE, "Meilenstein"),
        ("FF0000", "Heute"),
        (C_WEEKEND, "Wochenende"),
    ]
    ws.cell(row=legend_row, column=1).value = "Legende:"
    ws.cell(row=legend_row, column=1).font = Font(bold=True, size=9)

    for li, (color, label) in enumerate(legend_items):
        col = 2 + li * 2
        c_box = ws.cell(row=legend_row, column=col)
        c_box.fill = col_fill(color)
        c_box.value = "  "
        c_lbl = ws.cell(row=legend_row, column=col + 1)
        c_lbl.value = label
        c_lbl.font = Font(size=9)

    ws.freeze_panes = ws.cell(row=5, column=DATE_START)

    if not output_path:
        safe = project_name.replace(" ", "_")
        output_path = f"C:\\Users\\USass\\Billi_Bilder\\PM_Tool\\{safe}_Gantt.xlsx"

    wb.save(output_path)
    print(f"[OK] Gantt gespeichert: {output_path}")
    return output_path


if __name__ == "__main__":
    test_data = {
        "project_name": "PM Tool",
        "project_manager": "Udo Sassik",
        "tasks": [
            {
                "name": "Charter Generator",
                "owner": "Claude + OpenCode",
                "start": "2026-04-07",
                "end": "2026-04-09",
                "status": "Fertig",
            },
            {
                "name": "Canvas Generator",
                "owner": "Claude + OpenCode",
                "start": "2026-04-09",
                "end": "2026-04-10",
                "status": "Fertig",
            },
            {
                "name": "Gantt Generator",
                "owner": "Claude + OpenCode",
                "start": "2026-04-10",
                "end": "2026-04-11",
                "status": "In Arbeit",
            },
            {
                "name": "KI-Assistent (Ollama)",
                "owner": "Udo + Jules",
                "start": "2026-04-12",
                "end": "2026-04-20",
                "status": "Offen",
            },
            {
                "name": "main.py CLI",
                "owner": "OpenCode",
                "start": "2026-04-21",
                "end": "2026-04-25",
                "status": "Offen",
            },
            {
                "name": "PM Tool v1.0 Release",
                "owner": "Udo Sassik",
                "start": "2026-04-30",
                "end": "2026-04-30",
                "status": "Offen",
                "milestone": True,
            },
        ],
    }

    create_gantt(test_data)
